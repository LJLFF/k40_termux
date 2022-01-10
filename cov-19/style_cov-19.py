from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font, Alignment

wb = load_workbook('新冠数据.xlsx')
ws = wb['昨日新冠疫情数据']#默认选择第一张工作表

red_fill = PatternFill('solid', start_color='FF0000')


#给第六列从第二行开始填充颜色
#第一种获得某列的方法
#cells = ws.iter_rows(min_row=2, min_col=6, max_col=6)
#for i in cells:
    #i[0].fill=red_fill
#每行的内容是元组，每个元组只有一个单元格，所以取第一项的value，然后转为int型
  
#第二种获得某个区域,遍历单元格填充
for j in ws['D2':'D225']:
	j[0].fill=red_fill
for j in ws['F2':'F225']:
	j[0].fill=red_fill
wb.save('新冠数据.xlsx')

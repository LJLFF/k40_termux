from urllib import request
from parsel import Selector
from openpyxl.styles import PatternFill,Font, Alignment
			
url = 'https://www.worldometers.info/coronavirus/?continueFlag=0712ca94bb9b18139810671c3ee6ba4c'
handler = request.HTTPSHandler()
headers=("User-Agent","Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36 QIHU 360SE")

opener = request.build_opener(handler)
request.install_opener(opener)
opener.addheaders=[headers]
#构造一个请求
req = request.Request(url)

#打开请求得到响应
res = opener.open(req)
text = res.read().decode()
sec = Selector(text).xpath('//table[@id="main_table_countries_yesterday"]/tbody[1]/tr[position()>8]')
#sample =sec.get()

#用openpyxl写入excel
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet(title='昨日新冠疫情数据', index=0)
ws.append(['id','country','total_cases','new_cases','total_deaths','new_deaths','total_recovered','new_recovered','population'])

for li in sec:
    id = li.xpath('./td[1]/text()').get()
    
    #3种标签
    country = li.xpath('./td[2]/a[@class="mt_a"]/text()|./td[2]/span/text()|./td[2]/text()').get()
    total_cases =li.xpath('./td[3]/text()').get()
    new_cases = li.xpath('./td[4]/text()').get()
    total_deaths =li.xpath('./td[5]/text()').get()
    new_deaths =li.xpath('./td[6]/text()').get()
    total_recovered = li.xpath('./td[7]/text()').get()
    new_recovered =li.xpath('./td[8]/text()').get()
    #两种
    population =li.xpath('./td[15]/a/text()|./td[15]/text()').get()
    print(id,country,total_cases,new_cases,total_deaths,new_deaths,total_recovered,new_recovered,population)
    ws.append([id,country,total_cases,new_cases,total_deaths,new_deaths,total_recovered,new_recovered,population])
wb.save('新冠数据.xlsx')
